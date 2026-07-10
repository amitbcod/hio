import openpyxl
import json
from pathlib import Path
from openpyxl.worksheet.datavalidation import DataValidation

file_path = r"c:\wamp64\www\holidaysio\hio_transport_06_data_entry_sheet_2026_approved.xlsx"

def get_cell_type(cell):
    """Determine the data type of a cell"""
    if cell.data_type == 'f':
        return 'formula'
    elif cell.data_type == 'd':
        return 'date'
    elif cell.data_type == 'b':
        return 'boolean'
    elif cell.data_type == 'n':
        return 'number'
    elif cell.data_type == 's':
        return 'text'
    else:
        return 'general'

def analyze_sheet(ws, sheet_name):
    """Analyze a worksheet and extract field information"""
    
    fields = []
    
    if ws.max_row < 1:
        return fields
    
    # Get headers from first row
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            headers.append((col_idx, cell.value))
    
    # For each header, analyze the column
    for col_idx, header in headers:
        field_info = {
            'field_name': str(header),
            'column': col_idx,
            'field_type': None,
            'required': None,
            'default_value': None,
            'options': [],
            'validation_rules': [],
            'description': None,
            'sample_values': []
        }
        
        # Collect sample values to infer type
        sample_values = []
        cell_types_found = []
        
        for row_idx in range(2, min(ws.max_row + 1, 50)):  # Check first 50 rows
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                sample_values.append(str(cell.value)[:50])
                cell_types_found.append(get_cell_type(cell))
        
        field_info['sample_values'] = sample_values[:5]
        
        # Infer field type from cell types and content
        if cell_types_found:
            if 'date' in cell_types_found:
                field_info['field_type'] = 'date'
            elif 'boolean' in cell_types_found:
                field_info['field_type'] = 'checkbox'
            elif 'number' in cell_types_found:
                field_info['field_type'] = 'number'
            else:
                field_info['field_type'] = 'text'
        
        # Check for data validation (dropdowns, etc.)
        for dv in ws.data_validations.dataValidation:
            if dv.sqref:
                for cell_range in dv.sqref.ranges:
                    # Check if our column is in this range
                    if col_idx >= cell_range.min_col and col_idx <= cell_range.max_col:
                        if dv.type == 'list':
                            field_info['field_type'] = 'dropdown'
                            if dv.formula1:
                                # Extract list values
                                values_str = str(dv.formula1).strip('"')
                                options = [v.strip() for v in values_str.split(',')]
                                field_info['options'] = options
                        elif dv.type == 'textLength':
                            field_info['validation_rules'].append(f"Text length: {dv.operator} {dv.formula1}")
                        elif dv.type == 'whole' or dv.type == 'decimal':
                            field_info['field_type'] = 'number'
                            field_info['validation_rules'].append(f"Number {dv.operator}: {dv.formula1}")
        
        # Check if required (look for patterns in empty cells)
        empty_count = sum(1 for row_idx in range(2, min(ws.max_row + 1, 50)) 
                         if not ws.cell(row=row_idx, column=col_idx).value)
        total_rows = min(ws.max_row - 1, 49)
        if empty_count == 0 and total_rows > 0:
            field_info['required'] = True
        elif empty_count > 0:
            field_info['required'] = False
        
        fields.append(field_info)
    
    return fields

# Main execution
if not Path(file_path).exists():
    print(f"File not found: {file_path}")
else:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        print("=" * 70)
        print("ALL SHEET NAMES IN THE WORKBOOK:")
        print("=" * 70)
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"{i}. {sheet_name}")
        
        target_sheets = [
            '1_Transport_Basic',
            '2_Accounting_and_Transaction',
            '3_Policies_Rules',
            '4_Reservation_and_Communication'
        ]
        
        all_data = {}
        
        for sheet_name in target_sheets:
            print(f"\n{'=' * 70}")
            print(f"ANALYZING SHEET: {sheet_name}")
            print("=" * 70)
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                fields = analyze_sheet(ws, sheet_name)
                all_data[sheet_name] = fields
                
                print(f"Total fields: {len(fields)}\n")
                
                for field in fields:
                    print(f"  Field Name: {field['field_name']}")
                    print(f"  - Type: {field['field_type']}")
                    print(f"  - Required: {field['required']}")
                    if field['options']:
                        print(f"  - Options: {', '.join(field['options'])}")
                    if field['validation_rules']:
                        print(f"  - Validation: {', '.join(field['validation_rules'])}")
                    if field['sample_values']:
                        print(f"  - Sample values: {field['sample_values']}")
                    print()
            else:
                print(f"  [Sheet '{sheet_name}' not found in workbook]")
                all_data[sheet_name] = []
        
        # Output JSON
        print("\n" + "=" * 70)
        print("JSON OUTPUT FOR CONFIGURATION:")
        print("=" * 70)
        print(json.dumps(all_data, indent=2, default=str))
        
        wb.close()
        
    except Exception as e:
        print(f"Error: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
